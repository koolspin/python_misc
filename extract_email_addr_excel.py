import sys
import re
import argparse
from openpyxl import load_workbook
from openpyxl import Workbook


# This utility will iterate through each row and column in an Excel spreadsheet to extract email addresses.
# The extracted addresses will be written to a new spreadsheet
# Author: Colin Turner (koolspin)
# Dependencies: openpyxl

def validate_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("input", help="Path to the input xlsx file - containing email addresses")
    parser.add_argument("output", help="Path to the output xlsx file")
    args = parser.parse_args()
    if args.input is None or args.output is None:
        print("You must provide a path to both the input and output spreadsheets", file=sys.stderr)
        return False, None, None
    else:
        return True, args.input, args.output


def extract_email_addresses(in_file, out_file):
    new_wb = Workbook()
    new_ws = new_wb.active
    new_row = 1

    wb = load_workbook(in_file)
    sheet_collection = wb.sheetnames
    if len(sheet_collection) > 0:
        ws = wb[sheet_collection[0]]
        for row in ws.rows:
            for col in row:
                # print('Value of col: {0}'.format(col))
                if col is not None and col.value is not None and isinstance(col.value, str):
                    # match = re.search(r'[\w\.-]+@[\w\.-]+', d.value)
                    match = re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,6}\b', col.value)
                    if match is not None:
                        g = match.group(0)
                        if g is not None:
                            print('Matched email: [{0}]'.format(g))
                            new_ws.cell(row=new_row, column=1).value = g
                            new_row = new_row + 1

        new_wb.save(out_file)

res = validate_args()
if res[0]:
    extract_email_addresses(res[1], res[2])
    exit(0)
else:
    exit(1)

